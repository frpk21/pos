from datetime import datetime
from queue import Empty
from unicodedata import decimal
from django.shortcuts import render

from django.views import generic

from catalogos.models import Categoria, Movimientos_detalle, SubCategoria, Producto, Iva, Movimientos, Formulacion, Tipos_movimientos

from generales.models import Terceros, Profile

from django.db import connections

from collections import namedtuple

from catalogos.forms import CategoriaForm, SubCategoriaForm, ProductoForm, IvaForm, MovimientosEncForm, DetalleMovimientosFormSet, FormulacionEncForm, FormulacionDetForm, DetalleFormulacionFormSet

from django.urls import reverse_lazy, reverse

from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin

from generales.views import SinPrivilegios

from django.contrib.messages.views import SuccessMessageMixin

from django.http import JsonResponse

from django.http import HttpResponse, HttpResponseRedirect

from datetime import date

from datetime import datetime, timedelta

from django.db.models import Sum

from barcode.ean import EuropeanArticleNumber13
import barcode


def MenuView(request, *args, **kwargs):
    template_name="catalogos/menu.html"
    context={'hoy': date.today(), 'mes':date.today().month, 'ano': date.today().year}
   
    return render(request, template_name, context)

def MenuInvView(request, *args, **kwargs):
    template_name="catalogos/menu_inv.html"
    context={'hoy': date.today()}
   
    return render(request, template_name, context)


class CategoriaView(LoginRequiredMixin, generic.ListView):
    model = Categoria
    template_name = "catalogos/categorias.html"
    context_object_name = "obj"
    login_url='generales:login'
    
    def get_queryset(self):
        return Categoria.objects.filter(usuario=self.request.user)


class CategoriaNew(SuccessMessageMixin, LoginRequiredMixin, SinPrivilegios, generic.CreateView):
    permission_required='catalogos.add_categoria'
    model=Categoria
    template_name="catalogos/categorias_form.html"
    context_object_name='obj1'
    form_class=CategoriaForm
    success_url=reverse_lazy("catalogos:categoria_list")
    success_message='Categoría creada satisfactoriamente'
#    login_url='generales:login'
    def post(self, request, *args, **kwargs):
        form = CategoriaForm(request.POST)
        if form.is_valid():
            return self.form_valid(form)
        else:
            return self.form_invalid(form)

    def form_valid(self, form):
        self.object = form.save(commit=False)
        self.object.usuario = self.request.user
        self.object = form.save()
        
        return HttpResponseRedirect(reverse_lazy("catalogos:categoria_list"))


    def form_invalid(self, form, detalle_movimientos, tipor):
        self.object=form
        return self.render_to_response(
            self.get_context_data(
                form=form,
                detalle_movimientos=detalle_movimientos
            )
        )


class CategoriaEdit(LoginRequiredMixin, SinPrivilegios, generic.UpdateView):
    permission_required='catalogos.change_categoria'
    model=Categoria
    template_name="catalogos/categorias_form.html"
    context_object_name='obj2'
    form_class=CategoriaForm
    success_url=reverse_lazy("catalogos:categoria_list")
#    login_url='generales:login'


class CategoriaDel(LoginRequiredMixin, SinPrivilegios, generic.DeleteView):
    permission_required='catalogos.delete_categoria'
    model=Categoria
    template_name="catalogos/categorias_del.html"
    context_object_name='obj3'
    form_class=CategoriaForm
    success_url=reverse_lazy("catalogos:categoria_list")
#    login_url='generales:login'


class SubCategoriaView(LoginRequiredMixin, generic.ListView):
    model = SubCategoria
    template_name = "catalogos/subcategorias.html"
    context_object_name = "obj"
    login_url='generales:login'
    
    def get_queryset(self):
        return SubCategoria.objects.filter(categoria__usuario=self.request.user)

class SubCategoriaNew(SuccessMessageMixin, LoginRequiredMixin, SinPrivilegios, generic.CreateView):
    permission_required='catalogos.add_subcategoria'
    model=SubCategoria
    template_name="catalogos/subcategorias_form.html"
    context_object_name='obj1'
    form_class=SubCategoriaForm
    success_url=reverse_lazy("catalogos:subcategoria_list")
    success_message='Sub Categoría creada satisfactoriamente'

class SubCategoriaEdit(SuccessMessageMixin,SinPrivilegios, generic.UpdateView):
    permission_required='catalogos.change_subcategoria'
    model=SubCategoria
    template_name="catalogos/subcategorias_form.html"
    context_object_name='obj'
    form_class=SubCategoriaForm
    success_url=reverse_lazy("catalogos:subcategoria_list")
    success_message='Sub Categoría modificada satisfactoriamente'

class SubCategoriaDel(LoginRequiredMixin, SinPrivilegios, generic.DeleteView):
    permission_required='catalogos.delete_subcategoria'
    model=SubCategoria
    template_name="catalogos/subcategorias_del.html"
    context_object_name='obj'
    form_class=SubCategoriaForm
    success_url=reverse_lazy("catalogos:subcategoria_list")


class ProductoView(LoginRequiredMixin, generic.ListView):
    model = Producto
    template_name = "catalogos/producto_list.html"
    context_object_name = "obj"
    login_url='generales:login'

    def get_queryset(self):
        return Producto.objects.filter(usuario=self.request.user)
    
    
    
class CatalogoView(LoginRequiredMixin, generic.ListView):
    model = Producto
    template_name = "catalogos/catalogo.html"
    context_object_name = "obj"
    login_url='generales:login'

    def get_queryset(self):
        return Producto.objects.filter(usuario=self.request.user)
    
    

class MovimientosMercanciaView(generic.CreateView):
    #permission_required = 'movimientos.add_movimientos'
    model = Movimientos
    login_url = 'generales:login'
    template_name = 'catalogos/movimientosform.html'
    form_class = MovimientosEncForm
    success_url = reverse_lazy('generales:home')

    def get(self, request, *args, **kwargs):

        ctx = {'fecha': datetime.today(), 'tipo': kwargs["tipoe"], 'tercero': 0, 'ubicacion': 1, 'tipo_movimiento': 1}

        self.object = None

        form = MovimientosEncForm(initial=ctx)

        detalle_movimientos_formset = DetalleMovimientosFormSet()

        return self.render_to_response( 
            self.get_context_data(
                form=form,
                tipo=kwargs["tipoe"],
                opc_tipo_movimiento= Tipos_movimientos.objects.filter(tipo=kwargs["tipoe"],usuario=self.request.user),
                detalle_movimientos=detalle_movimientos_formset            
            )
        )

    def post(self, request, *args, **kwargs):
        form = MovimientosEncForm(request.POST)
        detalle_movimientos = DetalleMovimientosFormSet(request.POST)
        tipor = kwargs["tipoe"]
        #print("SSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSS")
        #print(form.errors)
        #print(detalle_movimientos.errors)
        if form.is_valid() and detalle_movimientos.is_valid():
            return self.form_valid(form, detalle_movimientos, tipor)
        else:
            return self.form_invalid(form, detalle_movimientos, tipor)

    def form_valid(self, form, detalle_movimientos, tipor):
        self.object = form.save(commit=False)
        profile = Profile.objects.filter(user=self.request.user.id).get()
        if tipor == 1:
            self.object.documento_no = profile.entrada + 1
            profile.entrada = profile.entrada + 1
        else:
            self.object.documento_no = profile.salida + 1
            profile.salida = profile.salida + 1
        profile.save()
        self.object.usuario = self.request.user
        self.object = form.save()
        detalle_movimientos.instance = self.object
        detalle_movimientos.save()
        for detalle in detalle_movimientos:
            if detalle.cleaned_data["codigo_de_barra"]:
                producto=Producto.objects.filter(codigo_de_barra=detalle.cleaned_data["codigo_de_barra"], usuario=self.request.user.id).get()
                if tipor == 1:
                    if float(detalle.cleaned_data["costo"]) > 0:
                        producto.existencia = producto.existencia + detalle.cleaned_data["cantidad"]
                        if producto.costo_unidad != int(detalle.cleaned_data["costo"]):
                            producto.costo_unidad = int(detalle.cleaned_data["costo"]) + int(producto.costo_unidad) / 2
                else:
                    producto.existencia = producto.existencia - detalle.cleaned_data["cantidad"]
                producto.save()
        #imprimir(self, form, detalle_movimientos, tipor, producto)
        
        return HttpResponseRedirect(reverse_lazy("catalogos:menu_inv"))
        #return JsonResponse(
        #    {
        #        'content': {
        #            'message': 'Generado Documento No. '+str(self.object.id)+' exitosamente.'
        #        }
        #    }
        #)


    def form_invalid(self, form, detalle_movimientos, tipor):
        self.object=form
        return self.render_to_response(
            self.get_context_data(
                form=form,
                detalle_movimientos=detalle_movimientos
            )
        )




class ProductoNew(SuccessMessageMixin, LoginRequiredMixin, SinPrivilegios, generic.CreateView):
    permission_required='catalogos.add_producto'
    model = Producto
    template_name="catalogos/productos_form.html"
    context_object_name = 'obj'
    form_class = ProductoForm
    success_message='Producto creado satisfactoriamente'

    def get(self, request, *args, **kwargs):
        self.object = None
        form_class = self.get_form_class()
        form = self.get_form(form_class)

        return self.render_to_response(
            self.get_context_data(
                form=form,
                retorna=kwargs["pk"]
            )
        )

    def post(self, request, *args, **kwargs):
        form_class = self.get_form_class()
        form = self.get_form(form_class)
        #print("XXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXXX")
        #print(form.errors)
        if form.is_valid():
            return self.form_valid(form,kwargs["pk"])
        else:
            return self.form_invalid(form)

    def form_valid(self, form, pk):
        self.object = form.save(commit=False)
        self.object.usuario = self.request.user
        self.object = form.save()
        if pk == 1:
            return HttpResponseRedirect(reverse_lazy("catalogos:productos_list"))
        else:
            return HttpResponseRedirect(reverse_lazy("catalogos:catalogo"))

    def form_invalid(self, form):
        self.object = form
        return self.render_to_response(
            self.get_context_data(
                form=form
            )
        )



class ProductoEdit(SuccessMessageMixin,SinPrivilegios, generic.UpdateView):
    permission_required = 'catalogos.change_producto'
    model = Producto
    template_name = "catalogos/productos_form.html"
    context_object_name='obj'
    form_class=ProductoForm
    success_url=reverse_lazy("catalogos:productos_list")
    success_message='Producto actualizado satisfactoriamente'

class ProductoDel(LoginRequiredMixin, SinPrivilegios, generic.DeleteView):
    permission_required='catalogos.delete_producto'
    model=Producto
    template_name="catalogos/producto_del.html"
    context_object_name='obj'
    form_class=ProductoForm
    success_url=reverse_lazy("catalogos:productos_list")

class TarifasIvaListView(LoginRequiredMixin, generic.ListView):
    model = Iva
    template_name = "catalogos/iva_list.html"
    context_object_name = "obj"
    login_url='generales:login'

class TarifaIvaNew(SuccessMessageMixin, LoginRequiredMixin, SinPrivilegios, generic.CreateView):
    permission_required='catalogos.add_iva'
    model = Iva
    template_name="catalogos/iva_form.html"
    context_object_name = 'obj'
    form_class = IvaForm
    success_url = reverse_lazy("catalogos:tarifas_iva")
    success_message='Producto creado satisfactoriamente'

class TarifaIvaEdit(SuccessMessageMixin,SinPrivilegios, generic.UpdateView):
    permission_required = 'catalogos.change_iva'
    model = Iva
    template_name = "catalogos/iva_form.html"
    context_object_name='obj'
    form_class=IvaForm
    success_url=reverse_lazy("catalogos:tarifas_iva")
    success_message='Producto actualizado satisfactoriamente'

class TarifaIvaDel(LoginRequiredMixin, SinPrivilegios, generic.DeleteView):
    permission_required='catalogos.delete_iva'
    model = Iva
    template_name="catalogos/iva_del.html"
    context_object_name='obj'
    form_class=IvaForm
    success_url=reverse_lazy("catalogos:tarifas_iva")


def get_ajaxSubcategoria(request, *args, **kwargs): 
    query = request.GET.get('q', None)
    if query: 
        terceros = SubCategoria.objects.filter(nombre__icontains=query, categoria__usuario=request.user).values("id","nombre") 
        terceros = list(terceros)
        return JsonResponse(terceros, safe=False) 
    else: 
        return JsonResponse(data={'success': False, 'errors': 'No encuentro resultados.'}) 


def get_ajaxTerceros(request, *args, **kwargs): 
    query = request.GET.get('q', None)
    if query: 
        terceros = Terceros.objects.filter(rzn_social__icontains=query, user=request.user).values("id","rzn_social") 
        terceros = list(terceros)
        return JsonResponse(terceros, safe=False) 
    else: 
        return JsonResponse(data={'success': False, 'errors': 'No encuentro resultados.'}) 



def get_ajaxBarcode(request, *args, **kwargs): 
    bar_code = request.GET.get('bar_code', None)
    if not bar_code:
        return JsonResponse(data={'nombre': '', 'errors': 'No encuentro producto.'})
    else:
        bar_code = Producto.objects.filter(codigo_de_barra=bar_code, usuario=request.user).last()
        if bar_code:
            return JsonResponse(data={"nombre": bar_code.nombre, "costo_unidad": int(bar_code.costo_unidad)}, safe=False)
        else: 
            return JsonResponse(data={'nombre': '', 'errors': 'No encuentro producto.'})




def get_ajaxProducto(request, *args, **kwargs): 
    query = request.GET.get('q', None)
    if query: 
        terceros = Producto.objects.filter(nombre__icontains=query, usuario=request.user).values("id","nombre") 
        terceros = list(terceros)
        return JsonResponse(terceros, safe=False) 
    else: 
        return JsonResponse(data={'success': False, 'errors': 'No encuentro resultados.'})
    
    

def get_additem(request, *args, **kwargs):
    detalle_movimientos_formset = DetalleMovimientosFormSet()
    detalle_movimientos_formset.extra += 1
    return JsonResponse(data={'detalle_movimientos':detalle_movimientos_formset})


def get_ajaxCantidad(request, *args, **kwargs): 
    cantidad = request.GET.get('cantidad', None)
    bar_code = request.GET.get('producto', None)
    tipo = int(request.GET.get('tipo', None))
    if not cantidad:
        return JsonResponse(data={'existencia': '', 'errors': 'No encuentro producto.'})
    else:
        if tipo > 1:
            existencia = Producto.objects.filter(codigo_de_barra=bar_code, usuario=request.user).last()
            if bar_code:
                if int(cantidad) > existencia.existencia:
                    return JsonResponse(data={"errors": "CANTIDAD NO PUEDE SER SUPERIOR A LA EXISTENCIA DEL PRODUCTO"}, safe=False)
                else:
                    return JsonResponse(data={"errors": ""}, safe=False)
            else: 
                return JsonResponse(data={'errors': 'No encuentro producto.'})
        else:
            return JsonResponse(data={"errors": ""}, safe=False)
        
      
#def get_ajaxInformeMovimientos1(request, *args, **kwargs):
#    ano = int(request.GET.get('periodo')[0:4])
#    mes = int(request.GET.get('periodo')[5:7])
#    if not ano:
#        return JsonResponse(data={'errors': 'No hay datos.'})
#    else:
#        return HttpResponseRedirect(reverse('catalogos:info_movimientos1_resul', kwargs={'ano': ano,'mes':mes}))




class InformeMovimientos1View(LoginRequiredMixin, generic.ListView):
    
    model = Movimientos
    template_name = "catalogos/info_movimientos1.html"
    context_object_name = "obj"
    login_url='generales:login'

    def get(self, request, *args, **kwargs):
        ano = int(request.GET.get('periodo')[0:4])
        mes = int(request.GET.get('periodo')[5:7])
        movimientos = Movimientos_detalle.objects.filter(movimiento__fecha__month=mes, movimiento__fecha__year=ano, movimiento__usuario=request.user)
        context = {}
        context['mes'] = mes
        context['ano'] = ano
        context['empresa'] = request.user.profile.empresa
        context['movimientos'] = movimientos
        self.object_list = context

        return self.render_to_response(
            self.get_context_data(
                context = context
            )
        )
        
        
#   ********  barcode imprimir      
def BarCodePrint(request,pk):
    codigo=pk
    if len(codigo) < 12:
        codigo = '0' * (12-len(codigo))
    codigo = codigo + pk
    template_name = 'catalogos/barcode_print.html'
    ean = EuropeanArticleNumber13(str(codigo), writer=ImageWriter())
    img=ean.save("media/barcode")
    producto = Producto.objects.filter(id=int(pk)).last()
    context={}
    context['producto'] = producto
    context['codigo'] = img

    return render(request, template_name, context)




class FormulacionView(SuccessMessageMixin, LoginRequiredMixin, SinPrivilegios, generic.CreateView):
    permission_required = 'formulacion.add_formulacion'
    model = Formulacion
    login_url = 'generales:login'
    template_name = 'catalogos/formulacionform.html'
    form_class = FormulacionEncForm
    success_url = reverse_lazy('generales:home')

    def get(self, request, *args, **kwargs):

        self.object = None

        form = FormulacionEncForm()

        detalle_movimientos_formset = DetalleFormulacionFormSet()

        return self.render_to_response( 
            self.get_context_data(
                form=form,
                detalle_movimientos=detalle_movimientos_formset            
            )
        )

    def post(self, request, *args, **kwargs):
        
        form = FormulacionEncForm(request.POST)
        detalle_movimientos = DetalleFormulacionFormSet(request.POST)
        #print("SSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSSS")
        #print(form.errors)
        #print(detalle_movimientos.errors)
        if form.is_valid() and detalle_movimientos.is_valid():
            return self.form_valid(form, detalle_movimientos)
        else:
            return self.form_invalid(form, detalle_movimientos)

    def form_valid(self, form, detalle_movimientos):
        self.object = form.save(commit=False)
        self.object.usuario = self.request.user
        self.object = form.save()
        detalle_movimientos.instance = self.object
        detalle_movimientos.save()

        return HttpResponseRedirect(reverse_lazy("generales:home"))

    def form_invalid(self, form, detalle_movimientos):
        self.object=form
        return self.render_to_response(
            self.get_context_data(
                form=form,
                detalle_movimientos=detalle_movimientos
            )
        )



def imprimirMovimiento(request, pk, tipo, documento_no):
    import io
    import os
    from django.conf import settings
    from reportlab.platypus import SimpleDocTemplate, Paragraph, TableStyle
    from reportlab.lib.enums import TA_JUSTIFY, TA_LEFT, TA_CENTER, TA_RIGHT
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib import colors  
    from reportlab.lib.pagesizes import letter, landscape, portrait
    from reportlab.platypus import Table
    from reportlab.lib.units import inch, mm
    from reportlab.platypus import Image, PageBreak, Paragraph, Spacer
    from django.core.mail import EmailMessage
    from io import StringIO
    from reportlab.pdfgen import canvas
    from reportlab_qrcode import QRCodeImage
    from reportlab.graphics.barcode import code128

    response = HttpResponse(content_type='application/pdf')  
    buffer = io.BytesIO()

     
    ordenes = []
    #logo_path = user.profile.foto.url
    #logo = os.path.join(settings.BASE_DIR, logo_path)
    #texto_path = "static/base/images/texto-inrai.jpg"
    #texto = os.path.join(settings.BASE_DIR, texto_path)
    #image = Image(logo, 3 * inch, .8 * inch)
    #image.hAlign = "LEFT"
    #image1 = Image(texto, 2.5 * inch, .8 * inch)
    #image1.hAlign = "LEFT"
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='Normal_CENTER', alignment=TA_CENTER))
    styles.add(ParagraphStyle(name='Pie',
                          alignment=TA_CENTER,
                          fontName='Helvetica',
                          fontSize=18,
                          textColor=colors.darkgray,
                          leading=8,
                          textTransform='uppercase',
                          wordWrap='LTR',
                          splitLongWords=True,
                          spaceShrinkage=0.05,
                          ))
    styles.add(ParagraphStyle(name='ejemplo',
                          parent=styles['Normal'],
                          fontName='Helvetica',
                          wordWrap='LTR',
                          alignment=TA_CENTER,
                          fontSize=12,
                          leading=13,
                          textColor=colors.black,
                          borderPadding=0,
                          leftIndent=0,
                          rightIndent=0,
                          spaceAfter=0,
                          spaceBefore=0,
                          splitLongWords=True,
                          spaceShrinkage=0.05,
                          ))
    mov_res = (Movimientos_detalle.objects.filter(movimiento__documento_no=documento_no, movimiento__usuario=request.user))
    total_doc = 0
    for i,item in enumerate(mov_res):
        movimiento = item
        no_doc = item.movimiento.documento_no
        producto =  item.producto
        codigo_de_barra = item.codigo_de_barra
        cantidad = item.cantidad
        costo = item.costo
        total = item.total 
        total_doc = total_doc + total       

    if tipo == 1:
        filename = "Entreda_{}.pdf".format(documento_no)
        titulo = "ENTRADA DE ALMACEN # {}".format(documento_no)
    else:
        filename = "Salid_{}.pdf".format(documento_no)
        titulo = "SALIDA DE ALMACEN # {}".format(documento_no)
    subtitulo = movimiento.movimiento.tipo_movimiento.nombre
    #qr = QRCodeImage(str(codigo_de_barra), size=30 * mm)
    #qr.hAlign = "RIGHT"
    


    t=Table(
        data=[
            ['','',titulo,'',''],
            ['','','',subtitulo,''],
            ['TERCERO', '',movimiento.movimiento.tercero.rzn_social,'FECHA',movimiento.movimiento.fecha.strftime('%d/%m/%Y')],
            ['CIUDAD', '',movimiento.movimiento.ubicacion.ciudad.nombre_ciudad,'',''],
            ['DIRECCION', '',movimiento.movimiento.tercero.direccion,'VALOR  ','${:,}'.format(total_doc)],
            ['TELEFONO', '',movimiento.movimiento.tercero.tel1,'', '']
        ],
        colWidths=[30,14,316,40,140],
        style=[
                ("FONT", (0,2), (0,2), "Helvetica", 7, 7),
                ("FONT", (0,3), (0,3), "Helvetica", 7, 7),
                ("FONT", (0,4), (0,4), "Helvetica", 7, 7),
                ("FONT", (0,5), (0,5), "Helvetica", 7, 7),
                ("FONT", (2,3), (2,3), "Helvetica-Bold", 7, 7),
                ("FONT", (2,4), (2,4), "Helvetica-Bold", 7, 7),
                ("FONT", (2,5), (2,5), "Helvetica-Bold", 7, 7),
                ("FONT", (2,0), (2,0), "Helvetica-Bold", 14, 14),
                ("FONT", (3,0), (3,0), "Helvetica-Bold", 11, 11),
                ("FONT", (3,1), (3,1), "Helvetica", 7, 7),
                ("FONT", (3,2), (3,4), "Helvetica", 7, 7),
                ("FONT", (3,5), (3,5), "Helvetica", 7, 7),
                ("FONT", (4,5), (4,5), "Helvetica", 9, 9),

                ('VALIGN',(3,2), (4,4),'MIDDLE'),
                ('ALIGN',(4,2),(4,4),'CENTRE'),
                ('VALIGN',(0,0), (1,0),'BOTTOM'),
                ('VALIGN',(3,0), (4,0),'MIDDLE'),
                ('VALIGN',(3,1), (4,1),'TOP'),
                ('VALIGN',(3,5), (3,5),'MIDDLE'),
                ('VALIGN',(4,5), (4,5),'MIDDLE'),
                ('ALIGN',(4,5),(4,5),'CENTRE'),
 
                ('ALIGN',(2,0),(2,0),'RIGHT'),
                
                ('LINEBELOW', (0,2),(2,4), .5, colors.gray),
                ('LINEBELOW', (3,2),(3,4), .5, colors.gray),
                #('LINEBELOW', (3,2),(3,4), .5, colors.gray),

                ('SPAN',(2,0),(4,0)),
                
                ('SPAN',(3,2),(3,3)),  #fecha
                ('SPAN',(4,2),(4,3)),  # valor fecha
                
                ('SPAN',(3,4),(3,5)),  # titulo valor
                ('SPAN',(4,4),(4,5)),  # valor total documento

                ('VALIGN',(0,2),(2,5),'MIDDLE'),
                ('BOX',(0,2),(2,5),.5,colors.gray),
                ('BOX',(3,2),(4,5),.5,colors.gray),
                ('BOX',(3,2),(4,5),.5,colors.gray),
                ('LINEBELOW', (4,2),(4,4), .5, colors.gray),
            ]
        )

    t.hAlign = "LEFT"
    ordenes.append(t)
    ordenes.append(Spacer(1, 5))
    if tipo == 1:
        headings0 = ('PRODUCTO', 'CODIGO DE BARRA', 'COSTO', 'CANTIDAD', 'TOTAL')
    else:
        headings0 = ('PRODUCTO', 'CODIGO DE BARRA', 'CANTIDAD')
    recibos2=[]
    t_cantidad = 0
    for lin, reg in enumerate(mov_res):
        if reg.cantidad is not None:
            t_cantidad = t_cantidad +  reg.cantidad
            barcode_value = str(reg.codigo_de_barra)
            barcode128 = code128.Code128(
                barcode_value,
                barHeight=15,
                barWidth=1,
                fontSize=10,
                humanReadable = False
            )
            if tipo == 1:
                recibos2.append([
                    reg.producto,
                    barcode128,
                    '${:,}'.format(reg.costo),
                    '{:,}'.format(reg.cantidad),
                    '${:,}'.format(reg.total)
                    ])
            else:
                recibos2.append([
                    reg.producto,
                    barcode128,
                    '{:,}'.format(reg.cantidad)
                    ])
    if tipo == 1:
        recibos2.append([
        'TOTAL',
        '',
        '',
        '{:,}'.format(t_cantidad),
        '${:,}'.format(total_doc)
        ])
    else:
        recibos2.append([
        'TOTAL',
        '',
        '{:,}'.format(t_cantidad)
        ])

    if tipo == 1:
        t0 = Table([headings0] + recibos2, colWidths=[180,180,60,60,60])
        t0.setStyle(TableStyle(
        [  
            ('GRID', (0, 0), (4, -1), 1, colors.gray),  
            ('LINEBELOW', (0, 0), (-1, 0), 1, colors.gray),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONT', (0, 0), (4, -1), "Helvetica", 8,8),
            ('FONTSIZE', (0, 0), (4, -1), 7),
            ('BACKGROUND', (0, 0), (4,0), colors.gray)  
        ]  
        ))
        t0.hAlign = "LEFT"
    else:
        t0 = Table([headings0] + recibos2, colWidths=[180,180,60])
        t0.setStyle(TableStyle(
        [  
            ('GRID', (0, 0), (2, -1), 1, colors.gray),  
            ('LINEBELOW', (0, 0), (-1, 0), 1, colors.gray),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONT', (0, 0), (2, -1), "Helvetica", 8,8),
            ('FONTSIZE', (0, 0), (2, -1), 7),
            ('BACKGROUND', (0, 0), (2,0), colors.gray)  
        ]  
        ))
        t0.hAlign = "CENTER"
    ordenes.append(t0)

    ordenes.append(Spacer(1, 5))
    icon_path = "/static/base/images/favicon.png"
    icon = os.path.join(settings.BASE_DIR, icon_path)
    doc = SimpleDocTemplate(buffer,
                pagesize=portrait(letter),
                rightMargin=40,
                leftMargin=50,
                topMargin=20,  
                bottomMargin=8,
                author="INRAI",
                title=titulo,
                icon=icon,
                )
    
    doc.build(ordenes)
    response.write(buffer.getvalue())
    pdf = buffer.getvalue()
    buffer.close()
    
    
    return response



def imprimirCatalogo(request):
    import io
    import os
    from django.conf import settings
    from reportlab.platypus import SimpleDocTemplate, Paragraph, TableStyle
    from reportlab.lib.enums import TA_JUSTIFY, TA_LEFT, TA_CENTER, TA_RIGHT
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib import colors  
    from reportlab.lib.pagesizes import letter, landscape, portrait
    from reportlab.platypus import Table
    from reportlab.lib.units import inch, mm
    from reportlab.platypus import Image, PageBreak, Paragraph, Spacer
    from django.core.mail import EmailMessage
    from io import StringIO
    from reportlab.pdfgen import canvas
    from reportlab_qrcode import QRCodeImage
    from reportlab.graphics.barcode import code128

    response = HttpResponse(content_type='application/pdf')  
    buffer = io.BytesIO()

     
    ordenes = []
    #logo_path = user.profile.foto.url
    #logo = os.path.join(settings.BASE_DIR, logo_path)
    #texto_path = "static/base/images/texto-inrai.jpg"
    #texto = os.path.join(settings.BASE_DIR, texto_path)
    #image = Image(logo, 3 * inch, .8 * inch)
    #image.hAlign = "LEFT"
    #image1 = Image(texto, 2.5 * inch, .8 * inch)
    #image1.hAlign = "LEFT"
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name='Normal_CENTER', alignment=TA_CENTER))
    styles.add(ParagraphStyle(name='Pie',
                          alignment=TA_CENTER,
                          fontName='Helvetica',
                          fontSize=18,
                          textColor=colors.darkgray,
                          leading=8,
                          textTransform='uppercase',
                          wordWrap='LTR',
                          splitLongWords=True,
                          spaceShrinkage=0.05,
                          ))
    styles.add(ParagraphStyle(name='ejemplo',
                          parent=styles['Normal'],
                          fontName='Helvetica',
                          wordWrap='LTR',
                          alignment=TA_CENTER,
                          fontSize=12,
                          leading=13,
                          textColor=colors.black,
                          borderPadding=0,
                          leftIndent=0,
                          rightIndent=0,
                          spaceAfter=0,
                          spaceBefore=0,
                          splitLongWords=True,
                          spaceShrinkage=0.05,
                          ))
    mov_res = (Producto.objects.filter(usuario=request.user).order_by('nombre','codigo_de_barra'))
    

    t=Table(
        data=[
            [request.user.profile.empresa],
            ["EXISTENCIAS DE ALMACEN"],
            [date.today().strftime('%d/%m/%Y')]
        ],
        colWidths=[540],
        style=[
                ("FONT", (0,0), (0,0), "Helvetica-Bold", 14, 14),
                ("FONT", (1,0), (1,0), "Helvetica-Bold", 12, 12),
                ("FONT", (2,0), (2,0), "Helvetica", 9, 9),

                ('VALIGN',(0,0), (2,2),'MIDDLE'),
                ('ALIGN',(0,0),(2,2),'CENTRE'),
            ]
        )

    t.hAlign = "LEFT"
    ordenes.append(t)
    ordenes.append(Spacer(1, 5))

    headings0 = ('NOMBRE', 'DESCRIPCION', 'UND', 'EXISTENCIA', 'PRECIO')
    recibos2=[]
    total_existencia=0
    total_precio = 0
    for lin, reg in enumerate(mov_res):
        if reg.existencia is not None:
            total_existencia = total_existencia +  reg.existencia
            total_precio = total_precio + reg.costo_unidad
            if reg.unidad_de_medida == 1:
                und = 'UNIDAD'
            elif  reg.unidad_de_medida == 2:
                und = 'Kg'
            elif  reg.unidad_de_medida == 3:
                und = 'Gr'
            elif  reg.unidad_de_medida == 4:
                und = 'Mlg'
            elif  reg.unidad_de_medida == 5:
                und = 'Mtro'
            elif  reg.unidad_de_medida == 6:
                und = 'CM'
            elif  reg.unidad_de_medida == 7:
                und = 'MILM'
            elif  reg.unidad_de_medida == 8:
                und = 'Lit'
            elif  reg.unidad_de_medida == 9:
                und = 'Mli'
            elif  reg.unidad_de_medida == 10:
                und = 'Cml'
            else:
                und = 'CM2'
            recibos2.append([
                reg.nombre,
                reg.descripcion,
                und,
                '{:,}'.format(reg.existencia),
                '{:,}'.format(reg.costo_unidad)
                ])
    recibos2.append([
        'TOTAL',
        '',
        '',
        '{:,}'.format(total_existencia),
        '${:,}'.format(total_precio)
        ])

    t0 = Table([headings0] + recibos2, colWidths=[180,180,60,60,60])
    t0.setStyle(TableStyle(
        [  
            ('GRID', (0, 0), (4, -1), 1, colors.gray),  
            ('LINEBELOW', (0, 0), (-1, 0), 1, colors.gray),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONT', (3, 1), (3,0), "Helvetica-Bold", 10,10),
            ('FONTSIZE', (0, 0), (4, -1), 7),
            ('FONTSIZE', (3, 0), (3, -1), 9),
            ('BACKGROUND', (0, 0), (4,0), colors.gray)  
        ]  
    ))
    t0.hAlign = "LEFT"
    ordenes.append(t0)

    ordenes.append(Spacer(1, 5))
    icon_path = "/static/base/images/favicon.png"
    icon = os.path.join(settings.BASE_DIR, icon_path)
    doc = SimpleDocTemplate(buffer,
                pagesize=portrait(letter),
                rightMargin=40,
                leftMargin=50,
                topMargin=20,  
                bottomMargin=8,
                author="INRAI",
                title="EXISTENCIAS DE ALMACEN VALORIZADO POR VENTA",
                icon=icon,
                )
    
    doc.build(ordenes)
    response.write(buffer.getvalue())
    pdf = buffer.getvalue()
    buffer.close()
    
    
    return response


def imprimirCatalogoXls(request):
    from openpyxl import Workbook
    recibos = Producto.objects.filter(usuario=request.user).order_by('nombre','codigo_de_barra')
    #Creamos el libro de trabajo
    wb = Workbook()
    #Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro
    ws = wb.active
    #En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
        ######ws['A1'] = 'REPORTE DE PERSONAS'
    #Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
        ######ws.merge_cells('B1:E1')

    #Creamos los encabezados desde la celda A1 hasta
    ws['A1'] = 'SUB CATEGORIA'
    ws['B1'] = 'NOMBRE'
    ws['C1'] = 'DESCRIPCION'
    ws['D1'] = 'UNIDAD'
    ws['E1'] = 'EXISTENCIA'
    ws['F1'] = 'COSTO UNIDAD'
    ws['G1'] = 'TARIFA IVA'
    ws['H1'] = 'PRECIO DE VENTA'
    ws['I1'] = 'CODIGO DE BARRA'
    ws['J1'] = 'UNICACION'
    cont=2
    #Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
    for item in recibos:
        ws.cell(row=cont,column=1).value = item.subcategoria.nombre
        ws.cell(row=cont,column=2).value = item.nombre
        ws.cell(row=cont,column=3).value = item.descripcion
        ws.cell(row=cont,column=4).value = item.unidad_de_medida
        ws.cell(row=cont,column=5).value = item.existencia
        ws.cell(row=cont,column=6).value = item.costo_unidad
        ws.cell(row=cont,column=7).value = str(item.tarifa_iva)
        ws.cell(row=cont,column=8).value = item.precio_de_venta
        ws.cell(row=cont,column=9).value = item.ubicacion.descripcion
        cont = cont + 1

    nombre_archivo = "inventario-"+date.today().strftime('%d/%m/%Y')+".xls"
    #Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
    response = HttpResponse(content_type="application/ms-excel") 
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)

    return response





  
    
def imprimir(self, form, detalle_movimientos, tipor, producto):
    import io
    import os
    doc_to_printer=str(14)+str(11)+str(500)
    doc_to_printer=str(doc_to_printer+".txt")
    escritura=open(doc_to_printer,"w")
    escritura.write(self.request.user.profile.empresa+"\n")
    escritura.write(self.request.user.profile.direccion+"\n")
    escritura.write(self.request.user.profile.telefono+"\n")
    escritura.write("========================================\n")
    if tipor == 1:
        doc_no = "ENTRADA # {}".format(self.object.documento_no)
    else:
        doc_no = "SALIDA # {}".format(self.object.documento_no)
    escritura.write(doc_no+"     "+self.object.fecha.strftime('%d/%m/%Y')+"\n")
    escritura.write("TERCERO: "+self.object.tercero.rzn_social+"\n")
    escritura.write("CIUDAD: "+self.object.ubicacion.ciudad.nombre_ciudad+"\n")
    escritura.write("PRODUCTO           CANTIDAD  COSTO  TOTAL"+"\n")
    escritura.write("----------------------------------------\n")
    t_total = 0
    for detalle in detalle_movimientos:
        t_total = t_total + (int(detalle.cleaned_data["cantidad"]) * int(detalle.cleaned_data["costo"]))
        escritura.write(
            producto.nombre+"   "+
            str(detalle.cleaned_data["cantidad"])+
            str('${:,}'.format(int(detalle.cleaned_data["cantidad"])))+
            str('${:,}'.format(int(detalle.cleaned_data["cantidad"]) * int(detalle.cleaned_data["costo"])))+"\n"
            )
    escritura.write("========================================\n")
    escritura.write("TOTAL          "+str('${:,}'.format(self.object.valor_documento))+"\n")
    escritura.close()
    os.startfile(doc_to_printer,"print")
    return